"""Spreadsheet tools - read, write, edit Excel/CSV files."""

import csv
import io
import os
from pathlib import Path
from typing import Optional


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa
        return True
    except ImportError:
        return False


def _has_xlsxwriter() -> bool:
    try:
        import xlsxwriter  # noqa
        return True
    except ImportError:
        return False


def read_csv(csv_path: str, delimiter: str = ",", has_header: bool = True, max_rows: Optional[int] = None) -> str:
    """Read a CSV file and return as formatted text/table.

    Args:
        csv_path: Path to CSV file
        delimiter: Field delimiter (default: comma)
        has_header: Whether the first row is a header
        max_rows: Maximum number of rows to read (None = all)

    Returns:
        Formatted text representation of the CSV
    """
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        return "Empty CSV file."

    if max_rows and len(rows) > max_rows + 1:
        rows = rows[:max_rows + 1]

    # Calculate column widths
    col_widths = []
    for col_idx in range(max(len(r) for r in rows)):
        widths = [len(str(r[col_idx])) for r in rows if col_idx < len(r)]
        col_widths.append(max(widths) if widths else 5)

    # Format as table
    lines = []
    for ri, row in enumerate(rows):
        cells = []
        for ci, cell in enumerate(row):
            if ci < len(col_widths):
                cells.append(str(cell).ljust(col_widths[ci]))
        lines.append(" | ".join(cells))
        if ri == 0 and has_header:
            lines.append("-+-".join("-" * w for w in col_widths))

    result = "\n".join(lines)
    result += f"\n\n({len(rows)} rows x {max(len(r) for r in rows)} columns)"
    if max_rows and len(rows) > max_rows:
        result += f" [showing first {max_rows} rows]"

    return result


def read_excel(excel_path: str, sheet_name: Optional[str] = None, max_rows: Optional[int] = None) -> str:
    """Read an Excel file and return as formatted text.

    Args:
        excel_path: Path to .xlsx/.xls file
        sheet_name: Specific sheet to read (None = first sheet)
        max_rows: Maximum rows to read per sheet

    Returns:
        Formatted text representation
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if not _has_openpyxl():
        # Try installing openpyxl
        raise ImportError(
            "openpyxl is required to read Excel files. "
            "Install it with: pip install openpyxl"
        )

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}")
        sheets = [wb[sheet_name]]
    else:
        sheets = [wb[wb.sheetnames[0]]]

    output_parts = []
    for sheet in sheets:
        lines = [f"=== Sheet: {sheet.title} ==="]
        rows = []
        for ri, row in enumerate(sheet.iter_rows(values_only=True)):
            if max_rows and ri >= max_rows:
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            lines.append("(empty sheet)")
            output_parts.append("\n".join(lines))
            continue

        # Calculate column widths
        col_widths = []
        for col_idx in range(max(len(r) for r in rows)):
            widths = [len(r[col_idx]) for r in rows if col_idx < len(r)]
            col_widths.append(max(widths) if widths else 5)

        for ri, row in enumerate(rows):
            cells = []
            for ci, cell in enumerate(row):
                if ci < len(col_widths):
                    cells.append(cell.ljust(col_widths[ci]))
            lines.append(" | ".join(cells))
            if ri == 0:
                lines.append("-+-".join("-" * w for w in col_widths))

        lines.append(f"({len(rows)} rows x {max(len(r) for r in rows)} columns)")
        output_parts.append("\n".join(lines))

    wb.close()
    return "\n\n".join(output_parts)


def write_csv(csv_path: str, data: str, delimiter: str = ",") -> str:
    """Write data to a CSV file.

    Args:
        csv_path: Path to write CSV file
        data: Data as text (rows separated by newlines, cells by delimiter)
        delimiter: Field delimiter (default: comma)

    Returns:
        Confirmation message
    """
    path = Path(csv_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    lines = data.strip().split("\n")
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for line in lines:
            if line.strip():
                writer.writerow([cell.strip() for cell in line.split(delimiter)])

    return f"CSV written: {csv_path} ({len([l for l in lines if l.strip()])} rows)"


def write_excel(excel_path: str, data: str, sheet_name: str = "Sheet1") -> str:
    """Write data to an Excel file.

    Args:
        excel_path: Path to write .xlsx file
        data: Data as text (rows by newlines, cells by tabs or pipes)
        sheet_name: Name of the sheet

    Returns:
        Confirmation message
    """
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if not _has_xlsxwriter():
        try:
            import openpyxl  # noqa
        except ImportError:
            raise ImportError(
                "xlsxwriter or openpyxl is required to write Excel files. "
                "Install with: pip install xlsxwriter"
            )

    import xlsxwriter
    wb = xlsxwriter.Workbook(str(path))
    ws = wb.add_worksheet(sheet_name)

    lines = data.strip().split("\n")
    row_count = 0
    for ri, line in enumerate(lines):
        if not line.strip():
            continue
        # Auto-detect delimiter: tab > pipe > comma
        if "\t" in line:
            cells = line.split("\t")
        elif " | " in line:
            cells = [c.strip() for c in line.split(" | ")]
        else:
            cells = [c.strip() for c in line.split(",")]
        for ci, cell in enumerate(cells):
            # Try to write as number if possible
            try:
                if "." in cell:
                    ws.write_number(ri, ci, float(cell))
                else:
                    ws.write_number(ri, ci, int(cell))
            except (ValueError, TypeError):
                ws.write(ri, ci, cell)
        row_count += 1

    wb.close()
    return f"Excel written: {excel_path} ({row_count} rows)"


def list_excel_sheets(excel_path: str) -> str:
    """List all sheet names in an Excel file."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if not _has_openpyxl():
        raise ImportError("openpyxl required. Install: pip install openpyxl")

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    lines = [f"Sheets in {path.name}:", ""]
    for s in sheets:
        lines.append(f"  - {s}")
    return "\n".join(lines)
